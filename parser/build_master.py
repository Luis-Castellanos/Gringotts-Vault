"""
Reference implementation for creating / appending to the master transactions xlsx.

USAGE FROM THE SKILL:
    Don't invoke this directly as a CLI. Read it, understand the patterns, then
    write a tailored version that includes the transactions you parsed from the
    PDFs. The Python below is meant to be lifted into your parsing script.

KEY FUNCTIONS:
    build_or_append(transactions, output_path, existing_path=None)
        - transactions: list of dicts with keys
            date (datetime.date), account (str), account_number (str),
            source (str), category (str), subcategory (str), amount (float),
            balance (float or None), stmt_period (str), source_file (str)
        - output_path: where to save the final xlsx
        - existing_path: if provided, append to this file instead of creating new

DUPLICATE DETECTION:
    A row is a duplicate if Date + Account + Source + Amount all match an existing
    row. Account is in the key so the same charge on two different cards is NOT
    treated as a duplicate. Account # is intentionally NOT in the key (Account
    label already disambiguates and we don't want blank account-numbers on
    legacy rows to break dedup). Duplicates are skipped silently.

CONVENTIONS:
    - Sheet name: 'Transactions'
    - Header row: row 1, bold, light gray fill, frozen
    - Excel Table named 'tblTransactions' over the data range
    - Date format: mm/dd/yyyy
    - Amount/Balance format: #,##0.00;(#,##0.00);-
    - Account # stored as TEXT (preserves leading zeros)
    - Sorted by Date ascending after every write

SCHEMA (10 columns):
    Date | Account | Account # | Source | Category | Sub-category | Amount | Balance | Stmt period | Source file

LEGACY SCHEMAS handled by automatic migration on append:
    v1 (7 cols, no Account / Source file)
    v2 (9 cols, no Account #) — Account # is auto-extracted from the trailing
        4-digit substring of the Account label if present.
"""

from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

SHEET_NAME = "Transactions"
TABLE_NAME = "tblTransactions"
HEADERS = ["Date", "Account", "Account #", "Source", "Category", "Sub-category", "Amount", "Balance", "Stmt period", "Source file"]
COLUMN_WIDTHS = {"A": 12, "B": 22, "C": 10, "D": 45, "E": 18, "F": 22, "G": 12, "H": 14, "I": 22, "J": 50}

HEADER_FILL = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
HEADER_FONT = Font(name="Arial", bold=True, size=11)
BODY_FONT = Font(name="Arial", size=11)


def _init_sheet(ws):
    """Set up headers, freeze pane, and column widths on a fresh sheet."""
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
    for col_letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width
    ws.freeze_panes = "A2"


def _write_row(ws, row_idx, txn):
    """Write one transaction row and apply formatting."""
    ws.cell(row=row_idx, column=1, value=txn["date"]).number_format = "mm/dd/yyyy"
    ws.cell(row=row_idx, column=2, value=txn.get("account", ""))
    # Account # written as TEXT to preserve leading zeros (e.g. "0042")
    acct_num_cell = ws.cell(row=row_idx, column=3, value=str(txn.get("account_number", "")))
    acct_num_cell.number_format = "@"
    ws.cell(row=row_idx, column=4, value=txn["source"])
    ws.cell(row=row_idx, column=5, value=txn["category"])
    ws.cell(row=row_idx, column=6, value=txn["subcategory"])
    amount_cell = ws.cell(row=row_idx, column=7, value=txn["amount"])
    amount_cell.number_format = "#,##0.00;(#,##0.00);-"
    bal = txn.get("balance")
    if bal is not None:
        bal_cell = ws.cell(row=row_idx, column=8, value=bal)
        bal_cell.number_format = "#,##0.00;(#,##0.00);-"
    ws.cell(row=row_idx, column=9, value=txn["stmt_period"])
    ws.cell(row=row_idx, column=10, value=txn.get("source_file", ""))
    for col_idx in range(1, len(HEADERS) + 1):
        ws.cell(row=row_idx, column=col_idx).font = BODY_FONT


def _read_existing_keys(ws):
    """Return a set of (date, account, source, amount) tuples for duplicate detection.

    Account is included so the same charge appearing on two different cards
    is NOT treated as a duplicate. Account # is intentionally NOT in the key
    because Account already serves as the disambiguator and we don't want
    a missing/blank account # to break dedup on legacy rows.
    """
    keys = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        d = row[0]
        if isinstance(d, datetime):
            d = d.date()
        # New schema: Date(0), Account(1), Account#(2), Source(3), ..., Amount(6)
        keys.add((
            d,
            str(row[1] or "").strip(),
            str(row[3] or "").strip(),
            round(float(row[6] or 0), 2),
        ))
    return keys


def _refresh_table(ws, last_row):
    """Add or replace the Excel Table covering A1:G{last_row}."""
    # Remove any existing table with our name
    existing = [name for name in ws.tables]
    for name in existing:
        del ws.tables[name]
    if last_row < 2:
        return
    ref = f"A1:{get_column_letter(len(HEADERS))}{last_row}"
    table = Table(displayName=TABLE_NAME, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def _sort_by_date(ws):
    """Sort rows 2..end by Date ascending. Reads, sorts, rewrites."""
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        rows.append(row)

    def sort_key(r):
        d = r[0]
        if isinstance(d, datetime):
            d = d.date()
        return d if isinstance(d, date) else date.min

    rows.sort(key=sort_key)

    # Clear data rows
    if ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row - 1)

    # Rewrite
    for i, row in enumerate(rows, start=2):
        for col_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=i, column=col_idx, value=val)
            cell.font = BODY_FONT
            if col_idx == 1:
                cell.number_format = "mm/dd/yyyy"
            elif col_idx == 3:  # Account # — preserve as text
                cell.number_format = "@"
            elif col_idx in (7, 8):  # Amount, Balance
                cell.number_format = "#,##0.00;(#,##0.00);-"


def _migrate_old_schema(ws):
    """
    Detect and migrate older schemas to the current 10-column schema.

    v1 (7 cols): Date | Source | Category | Sub-category | Amount | Balance | Stmt period
    v2 (9 cols): Date | Account | Source | Category | Sub-category | Amount | Balance | Stmt period | Source file
    v3 (10 cols, current):
                 Date | Account | Account # | Source | Category | Sub-category | Amount | Balance | Stmt period | Source file

    Returns True if a migration happened, False if the sheet was already current
    (or empty / unrecognizable).
    """
    if ws.max_row < 1:
        return False
    headers = [c.value for c in ws[1]]
    if headers == HEADERS:
        return False

    v1_headers = ["Date", "Source", "Category", "Sub-category", "Amount", "Balance", "Stmt period"]
    v2_headers = ["Date", "Account", "Source", "Category", "Sub-category", "Amount", "Balance", "Stmt period", "Source file"]

    if headers == v1_headers:
        legacy_version = "v1"
    elif headers == v2_headers:
        legacy_version = "v2"
    else:
        return False  # Unknown schema, don't try to migrate

    # Read all old rows
    old_rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        old_rows.append(row)

    # Wipe sheet and rewrite with new headers
    ws.delete_rows(1, ws.max_row)
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
    for col_letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width
    ws.freeze_panes = "A2"

    # Re-write old rows mapped onto the new schema
    for i, old in enumerate(old_rows, start=2):
        if legacy_version == "v1":
            # v1: (date, source, cat, subcat, amt, bal, stmt)
            d, src, cat, subcat, amt, bal, stmt = old
            account = ""  # legacy: blank
            account_num = ""
            source_file = ""
        else:  # v2
            # v2: (date, account, source, cat, subcat, amt, bal, stmt, source_file)
            d, account, src, cat, subcat, amt, bal, stmt, source_file = old
            account_num = ""  # v2 didn't have account #; try to extract from account label
            # If account label ends with 4 digits (e.g. "Chase Checking 1234"), pull them out
            import re as _re
            m_num = _re.search(r"\b(\d{4})\s*$", str(account or ""))
            if m_num:
                account_num = m_num.group(1)

        ws.cell(row=i, column=1, value=d).number_format = "mm/dd/yyyy"
        ws.cell(row=i, column=2, value=account)
        an_cell = ws.cell(row=i, column=3, value=str(account_num))
        an_cell.number_format = "@"
        ws.cell(row=i, column=4, value=src)
        ws.cell(row=i, column=5, value=cat)
        ws.cell(row=i, column=6, value=subcat)
        ac = ws.cell(row=i, column=7, value=amt)
        ac.number_format = "#,##0.00;(#,##0.00);-"
        if bal is not None:
            bc = ws.cell(row=i, column=8, value=bal)
            bc.number_format = "#,##0.00;(#,##0.00);-"
        ws.cell(row=i, column=9, value=stmt)
        ws.cell(row=i, column=10, value=source_file)
        for col_idx in range(1, len(HEADERS) + 1):
            ws.cell(row=i, column=col_idx).font = BODY_FONT
    return True


def build_or_append(transactions, output_path, existing_path=None):
    """
    Build a new master xlsx or append to an existing one.

    Returns dict with keys: written, skipped_duplicates, total_rows
    """
    output_path = Path(output_path)

    if existing_path and Path(existing_path).exists():
        wb = load_workbook(existing_path)
        if SHEET_NAME not in wb.sheetnames:
            ws = wb.create_sheet(SHEET_NAME)
            _init_sheet(ws)
        else:
            ws = wb[SHEET_NAME]
            # Migrate old 7-col schema if needed
            _migrate_old_schema(ws)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        _init_sheet(ws)

    existing_keys = _read_existing_keys(ws)
    next_row = ws.max_row + 1 if ws.max_row >= 1 else 2
    if next_row < 2:
        next_row = 2

    written = 0
    skipped = 0
    for txn in transactions:
        key = (
            txn["date"],
            str(txn.get("account", "")).strip(),
            str(txn["source"]).strip(),
            round(float(txn["amount"]), 2),
        )
        if key in existing_keys:
            skipped += 1
            continue
        _write_row(ws, next_row, txn)
        existing_keys.add(key)
        next_row += 1
        written += 1

    _sort_by_date(ws)
    last_row = ws.max_row
    _refresh_table(ws, last_row)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    return {
        "written": written,
        "skipped_duplicates": skipped,
        "total_rows": last_row - 1,
    }


# ---- Quick self-test (run this file directly to verify it works) ----
if __name__ == "__main__":
    sample = [
        {
            "date": date(2024, 3, 15),
            "account": "Chase Checking 1234",
            "account_number": "1234",
            "source": "STARBUCKS STORE 1234",
            "category": "Food & Dining",
            "subcategory": "Coffee & Tea",
            "amount": -6.45,
            "balance": 1234.56,
            "stmt_period": "03/01/2024 - 03/31/2024",
            "source_file": "Chase_Checking_1234_03012024.pdf",
        },
        {
            "date": date(2024, 3, 16),
            "account": "Chase Checking 1234",
            "account_number": "1234",
            "source": "PAYROLL DEPOSIT EMPLOYER",
            "category": "Income",
            "subcategory": "Paycheck",
            "amount": 2500.00,
            "balance": 3734.56,
            "stmt_period": "03/01/2024 - 03/31/2024",
            "source_file": "Chase_Checking_1234_03012024.pdf",
        },
        {
            "date": date(2024, 3, 16),
            "account": "Chase Checking 1234",
            "account_number": "1234",
            "source": "PAYROLL DEPOSIT EMPLOYER",
            "category": "Income",
            "subcategory": "Paycheck",
            "amount": 2500.00,  # Duplicate of previous (same account)
            "balance": 3734.56,
            "stmt_period": "03/01/2024 - 03/31/2024",
            "source_file": "Chase_Checking_1234_03012024.pdf",
        },
        {
            "date": date(2024, 3, 16),
            "account": "Apple Card 5678",
            "account_number": "5678",  # DIFFERENT account → not a dupe
            "source": "PAYROLL DEPOSIT EMPLOYER",
            "category": "Income",
            "subcategory": "Paycheck",
            "amount": 2500.00,
            "balance": None,
            "stmt_period": "03/01/2024 - 03/31/2024",
            "source_file": "Apple_Card_5678_03012024.pdf",
        },
    ]
    result = build_or_append(sample, "/tmp/test_master.xlsx")
    print(f"Self-test result: {result}")
    assert result["written"] == 3, f"Expected 3 written, got {result['written']}"
    assert result["skipped_duplicates"] == 1, f"Expected 1 dupe, got {result['skipped_duplicates']}"
    print("Self-test passed.")
